from openpyxl import load_workbook
import csv
from datetime import datetime


def write_csv(data):
    with open('cms.xlsx', 'a') as f:
        writer = csv.writer(f)
        writer.writerow((data['LegalName'],
                         data['Email'],
                         data['ContactName']))



def refactoring(field):
    ans = ''
    for i in str(field).split():
        ans += str(i[0].upper() + i[1::].lower()) + ' '
        continue
    return ans



def search():
    startTime = datetime.now()
    counter = 49
    for i in range(48):
        workbook = load_workbook(filename=f"data/data{counter}.xlsx")
        counter += 1
        sheet = workbook.active
        amount = sheet.max_row
        list_mails = ['@gmail.com', '@yahoo.com', '@yahoo.com.mx', '@hotmail.com', '@icloud.com', '@outlook.com',
                      '@live.com', '@aol.com', '@usa.com', '@ymail.com']


        for i in range(2, amount):
             email = (sheet.cell(row=i, column=13).value)
             legalname = (sheet.cell(row=i, column=4).value)
             contact_name = (sheet.cell(row=i, column=14).value)

             if email == None:
                 pass
             else:
                 data = {'LegalName': refactoring(legalname),
                          'Email': email,
                          'ContactName': refactoring(contact_name)}

                 write_csv(data)
    print(counter)




             # for domain in list_mails:
             #     if email == None:
             #         pass
             #     elif domain in email:
             #         data = {'LegalName': refactoring(legalname),
             #                 'Email': email,
             #                 'ContactName': refactoring(contact_name)}
             #         write_csv(data)
             #     else:
             #         pass
             #     continue
    print(datetime.now() - startTime)





search()
